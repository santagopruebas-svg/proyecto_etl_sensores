# src/extract.py
import os
from openpyxl import load_workbook
from src.config import obtener_configuracion_por_nombre_interno, obtener_celda_pasillo

def encontrar_archivos_por_procesar(input_folder):
    """Busca archivos .xlsx en la carpeta de entrada especificada."""
    archivos_a_procesar = []
    if not os.path.exists(input_folder):
        print(f"ERROR: La carpeta de entrada no existe: {input_folder}")
        return archivos_a_procesar

    archivos_en_input = [f for f in os.listdir(input_folder) if f.endswith('.xlsx') and not f.startswith('~$')]
    # Nota: Agregué 'not f.startswith('~$')' para ignorar archivos temporales de Excel abiertos

    for filename in archivos_en_input:
        filepath = os.path.join(input_folder, filename)
        archivos_a_procesar.append(filepath)
            
    return archivos_a_procesar

def leer_archivo_excel(filepath):
    """
    Lee datos de un archivo Excel y CIERRA la conexión para permitir moverlo después.
    """
    filename = os.path.basename(filepath)
    # print(f"--- Leido: {filename} ---") # Opcional: comentar para menos ruido
    
    workbook = None
    try:
        # Cargar el workbook
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        sheet = workbook.active
        
        # 1. Identificar Pasillo
        celda_pasillo_key = obtener_celda_pasillo(filename)
        nombre_interno_pasillo = sheet[celda_pasillo_key].value
        
        if not nombre_interno_pasillo:
            return None, None, None

        # 2. Configuración
        config = obtener_configuracion_por_nombre_interno(str(nombre_interno_pasillo))
        
        if not config:
            return None, None, None

        data_start_row = config.get('data_start_row', 1) 
        
        # 3. Extraer Datos
        headers = []
        data_rows = []
        
        # Leer Cabeceras
        for row in sheet.iter_rows(min_row=data_start_row, max_row=data_start_row, values_only=True):
            headers = [str(cell).strip() if cell is not None else f"Col_{i}" for i, cell in enumerate(row)]
            break

        # Leer Datos
        start_data_row = data_start_row + 1
        for row in sheet.iter_rows(min_row=start_data_row, values_only=True):
            data_rows.append(list(row)) 
        
        return headers, data_rows, config
        
    except FileNotFoundError:
        print(f"ERROR: Archivo no encontrado: {filepath}")
        return None, None, None
    except Exception as e:
        print(f"ERROR al leer {filename}: {e}")
        return None, None, None
    finally:
        # --- BLOQUE CRÍTICO: CERRAR EL ARCHIVO ---
        if workbook:
            workbook.close()